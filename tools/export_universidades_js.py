#!/usr/bin/env python
"""Export university tracker workbooks to the static universidades.js format."""

from __future__ import annotations

import argparse
import datetime as dt
import json
import math
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from urllib.parse import quote

from openpyxl import load_workbook


BASE_SHEETS = ("Top 200 QS 2026", "Recomendadas AR-LATAM")
TRACKER_SHEETS = {
    "institutions": "Instituciones",
    "rankings": "Rankings_Historicos",
    "metrics": "Metricas_OpenAlex",
    "quality": "Control_Calidad",
}


ALIASES = {
    "n": ("n", "nro", "no", "numero", "num", "prioridad", "fila n en hoja principal"),
    "rank": ("ranking qs 2026", "ranking_base", "ranking base", "rank", "ranking", "ranking qs"),
    "name": ("universidad", "universidad_original", "institucion", "institution", "name", "universidad original"),
    "country": ("pais territorio qs", "pais territorio", "pais_original", "pais", "country"),
    "score": ("puntaje qs", "score", "puntaje"),
    "lat": ("latitud", "lat", "openalex_latitud", "ror_latitud"),
    "lng": ("longitud", "lng", "lon", "long", "openalex_longitud", "ror_longitud"),
    "maps": ("google maps", "maps"),
    "source": ("fuente ranking", "source", "fuente"),
}


def norm(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = re.sub(r"[°/#().,;:]+", " ", text)
    text = re.sub(r"\s+", " ", text)
    return (
        text.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value).strip()


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isnan(float(value)):
            return None
        return float(value)
    text = clean(value).replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return None


def find_header(rows: List[Tuple[Any, ...]]) -> Optional[int]:
    for i, row in enumerate(rows[:12]):
        names = {norm(c) for c in row if c is not None}
        has_name = any(a in names for a in ALIASES["name"])
        has_lat = any(a in names for a in ALIASES["lat"])
        has_lng = any(a in names for a in ALIASES["lng"])
        if has_name and (has_lat or has_lng):
            return i
    return None


def header_map(header: Iterable[Any]) -> Dict[str, int]:
    return {norm(value): idx for idx, value in enumerate(header) if value is not None}


def pick(row: Tuple[Any, ...], columns: Dict[str, int], key: str) -> Any:
    for alias in ALIASES[key]:
        idx = columns.get(norm(alias))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def maps_url(lat: float, lng: float) -> str:
    return "https://www.google.com/maps/search/?api=1&query=" + quote(f"{lat:.6f}, {lng:.6f}")


def row_to_record(row: Tuple[Any, ...], columns: Dict[str, int], fallback_n: int) -> Optional[Dict[str, Any]]:
    name = clean(pick(row, columns, "name"))
    lat = to_float(pick(row, columns, "lat"))
    lng = to_float(pick(row, columns, "lng"))
    if not name or lat is None or lng is None:
        return None

    raw_n = pick(row, columns, "n")
    n = int(to_float(raw_n) or fallback_n)
    rank = clean(pick(row, columns, "rank"))
    score = pick(row, columns, "score")
    score_text = "" if score is None else clean(score)

    return {
        "n": n,
        "rank": rank,
        "name": name,
        "country": clean(pick(row, columns, "country")),
        "score": score_text,
        "lat": round(lat, 6),
        "lng": round(lng, 6),
        "maps": clean(pick(row, columns, "maps")) or maps_url(lat, lng),
        "source": clean(pick(row, columns, "source")),
    }


def read_table_sheet(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    header_idx = find_header(rows)
    if header_idx is None:
        return []
    columns = header_map(rows[header_idx])
    out: List[Dict[str, Any]] = []
    for fallback_n, row in enumerate(rows[header_idx + 1 :], start=1):
        rec = row_to_record(row, columns, fallback_n)
        if rec:
            out.append(rec)
    return out


def key_for(name: str, country: str = "") -> str:
    return norm(f"{name} {country}")


def read_base_workbook(path: Path) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    by_key: Dict[str, Dict[str, Any]] = {}
    for sheet_name in BASE_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        for rec in read_table_sheet(wb[sheet_name]):
            by_key.setdefault(key_for(rec["name"], rec.get("country", "")), rec)
    return sorted(by_key.values(), key=lambda r: (r.get("n") or 999999, r.get("name") or ""))


def read_named_table(path: Path, sheet_name: str) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    return read_table_sheet(wb[sheet_name])


def read_generic_sheet(ws) -> List[Dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header_idx = None
    for i, row in enumerate(rows[:8]):
        filled = [c for c in row if c is not None]
        if len(filled) >= 3:
            header_idx = i
            break
    if header_idx is None:
        return []

    headers = [norm(c).replace(" ", "_") for c in rows[header_idx]]
    out: List[Dict[str, Any]] = []
    for row in rows[header_idx + 1 :]:
        rec: Dict[str, Any] = {}
        for idx, key in enumerate(headers):
            if not key or idx >= len(row):
                continue
            value = row[idx]
            if value is None:
                continue
            rec[key] = value
        if rec:
            out.append(rec)
    return out


def enrich_from_tracker(records: List[Dict[str, Any]], path: Path, target_year: int) -> List[Dict[str, Any]]:
    wb = load_workbook(path, data_only=True)
    if TRACKER_SHEETS["institutions"] not in wb.sheetnames:
        return records

    institutions = read_generic_sheet(wb[TRACKER_SHEETS["institutions"]])
    metrics = read_generic_sheet(wb[TRACKER_SHEETS["metrics"]]) if TRACKER_SHEETS["metrics"] in wb.sheetnames else []

    inst_by_key = {key_for(r.get("universidad_original", ""), r.get("pais_original", "")): r for r in institutions}
    met_by_key = {key_for(r.get("universidad", ""), r.get("pais", "")): r for r in metrics}

    for rec in records:
        k = key_for(rec.get("name", ""), rec.get("country", ""))
        inst = inst_by_key.get(k)
        if inst:
            for field in ("ror_id", "openalex_id", "homepage", "match_status"):
                if inst.get(field):
                    rec[field] = inst[field]
            lat = to_float(inst.get("latitud"))
            lng = to_float(inst.get("longitud"))
            if lat is not None and lng is not None:
                rec["tracker_lat"] = round(lat, 6)
                rec["tracker_lng"] = round(lng, 6)

        met = met_by_key.get(k)
        if met:
            for source, dest in (
                ("works_total", "works_total"),
                ("citas_total", "citas_total"),
                ("h_index", "h_index"),
                ("i10_index", "i10_index"),
                ("mean_citedness_2yr", "mean_citedness_2yr"),
                (f"publicaciones_{target_year}", f"publicaciones_{target_year}"),
                (f"citas_{target_year}", f"citas_{target_year}"),
                ("openalex_url", "openalex_url"),
            ):
                if met.get(source) not in (None, ""):
                    rec[dest] = met[source]
    return records


def js_literal(records: List[Dict[str, Any]], input_path: Path) -> str:
    payload = json.dumps(records, ensure_ascii=False, separators=(",", ":"))
    stamp = dt.datetime.now(dt.timezone.utc).replace(microsecond=0).isoformat()
    return (
        "/* Universidades generadas desde tracker/base Excel.\n"
        f"   Fuente: {input_path.as_posix()}\n"
        f"   Actualizado: {stamp}\n"
        "*/\n"
        f"window.UNIVERSIDADES = {payload};\n"
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export Excel university data to universidades.js")
    parser.add_argument("--input", required=True, help="Excel base or tracker-enriched workbook.")
    parser.add_argument("--output", default="universidades.js", help="JS file to write.")
    parser.add_argument("--target-year", type=int, default=dt.date.today().year - 1)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    input_path = Path(args.input).resolve()
    output_path = Path(args.output).resolve()
    records = read_base_workbook(input_path)

    if TRACKER_SHEETS["institutions"] in load_workbook(input_path, read_only=True).sheetnames:
        records = enrich_from_tracker(records, input_path, args.target_year)

    if not records:
        raise SystemExit(f"No university records found in {input_path}")

    output_path.write_text(js_literal(records, input_path), encoding="utf-8")
    print(f"Wrote {len(records)} universities to {output_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
